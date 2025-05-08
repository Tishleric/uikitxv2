"""
pMoneyAuto - Multi-Option Workflow using openpyxl for Excel

This version is adapted for integration with a dashboard.
It accepts option data as an argument, processes them,
returns a LIST of joined DataFrames, one for each input option,
AND writes this list of DataFrames to "Sheet1" of the Excel file.
"""

import time
import os
import io
import pandas as pd
import pyperclip
import webbrowser
import traceback
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows # For writing DFs to Excel
from pywinauto.keyboard import send_keys
import logging

logger = logging.getLogger(__name__) 

# --- Configuration Constants ---
FILE_PATH = r"\\ERIC-HOST-PC\FRGMSharedSpace\PnL Scanrios - US Rates.4.30.25.xlsx"
SHEET_NAME_PM_SETUP = "Sheet2" # Sheet used for PM interaction
PNL_SCENARIO_BUY_SHEET_NAME = "PnL Scenario - Buy"
PRICING_MONKEY_URL_FOR_PASTE = "https://pricingmonkey.com/b/3580a62f-daf9-49bd-bf2e-01120ff59371"

# New constants for Sheet1 output
TARGET_SHEET_FOR_OUTPUT = "Sheet1"
START_ROW_SHEET1 = 5
ROW_OFFSET_PER_DF_SHEET1 = 15 # Start next DF 15 rows below the start of the previous one

PHASE_CONFIGS = {
    1: {"excel_start_row": 3,  "excel_results_data_start_row": 3},
    2: {"excel_start_row": 6,  "excel_results_data_start_row": 6},
    3: {"excel_start_row": 8,  "excel_results_data_start_row": 8},
    4: {"excel_start_row": 10, "excel_results_data_start_row": 10},
    5: {"excel_start_row": 12, "excel_results_data_start_row": 12},
}
EXCEL_PM_DATA_END_ROW = 14
EXCEL_PM_RESULTS_MAX_ROW_FOR_CLEARING_UNUSED = 14
PNL_VALUE_TARGET_START_ROW_SHEET2 = 3
PNL_VALUE_TARGET_END_ROW_SHEET2 = 14

OPT1_USER_QTY_COL = "F"
OPT1_PM_QTY_COL = "G"
OPT1_PM_DESC_COL = "H"
OPT1_PM_COPY_END_COL = "I"
COL_OFFSET_FOR_SETUP_BLOCKS = 9

OPT2_USER_QTY_COL = get_column_letter(column_index_from_string(OPT1_USER_QTY_COL) + COL_OFFSET_FOR_SETUP_BLOCKS)
OPT2_PM_QTY_COL = get_column_letter(column_index_from_string(OPT1_PM_QTY_COL) + COL_OFFSET_FOR_SETUP_BLOCKS)
OPT2_PM_DESC_COL = get_column_letter(column_index_from_string(OPT1_PM_DESC_COL) + COL_OFFSET_FOR_SETUP_BLOCKS)
OPT2_PM_COPY_END_COL = get_column_letter(column_index_from_string(OPT1_PM_COPY_END_COL) + COL_OFFSET_FOR_SETUP_BLOCKS)

OPT3_USER_QTY_COL = get_column_letter(column_index_from_string(OPT2_USER_QTY_COL) + COL_OFFSET_FOR_SETUP_BLOCKS)
OPT3_PM_QTY_COL = get_column_letter(column_index_from_string(OPT2_PM_QTY_COL) + COL_OFFSET_FOR_SETUP_BLOCKS)
OPT3_PM_DESC_COL = get_column_letter(column_index_from_string(OPT2_PM_DESC_COL) + COL_OFFSET_FOR_SETUP_BLOCKS)
OPT3_PM_COPY_END_COL = get_column_letter(column_index_from_string(OPT2_PM_COPY_END_COL) + COL_OFFSET_FOR_SETUP_BLOCKS)

OPT1_PM_RESULT_START_COL = "J"
OPT1_PM_RESULT_END_COL = "L"
OPT2_PM_RESULT_START_COL = "S"
OPT2_PM_RESULT_END_COL = "U"
OPT3_PM_RESULT_START_COL = "AB"
OPT3_PM_RESULT_END_COL = "AD"

OPTION_COLUMN_MAP = [
    {"user_qty_col": OPT1_USER_QTY_COL, "pm_qty_col": OPT1_PM_QTY_COL, "pm_desc_col": OPT1_PM_DESC_COL, "pm_copy_end_col": OPT1_PM_COPY_END_COL,
     "pm_result_start_col": OPT1_PM_RESULT_START_COL, "pm_result_end_col": OPT1_PM_RESULT_END_COL},
    {"user_qty_col": OPT2_USER_QTY_COL, "pm_qty_col": OPT2_PM_QTY_COL, "pm_desc_col": OPT2_PM_DESC_COL, "pm_copy_end_col": OPT2_PM_COPY_END_COL,
     "pm_result_start_col": OPT2_PM_RESULT_START_COL, "pm_result_end_col": OPT2_PM_RESULT_END_COL},
    {"user_qty_col": OPT3_USER_QTY_COL, "pm_qty_col": OPT3_PM_QTY_COL, "pm_desc_col": OPT3_PM_DESC_COL, "pm_copy_end_col": OPT3_PM_COPY_END_COL,
     "pm_result_start_col": OPT3_PM_RESULT_START_COL, "pm_result_end_col": OPT3_PM_RESULT_END_COL},
]

PNL_BUY_SOURCE_COL = 'U'
PNL_BUY_SOURCE_ROWS = [4, 6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26]

KEY_PRESS_PAUSE = 0.01
WAIT_FOR_BROWSER_TO_OPEN = 2.5
WAIT_AFTER_BROWSER_NAVIGATION = 0.1
WAIT_AFTER_BROWSER_PASTE = 5.0
WAIT_FOR_COPY_OPERATION = 0.2
WAIT_FOR_BROWSER_CLOSE = 0.5

def check_file_exists(file_path):
    if os.path.exists(file_path): logger.info(f"Excel file verified at: {file_path}"); return True
    else: logger.error(f"Excel file not found at {file_path}."); return False

def excel_setup_options_and_consolidate_for_pm_openpyxl(file_path, sheet_name_target, pnl_sheet_name_source, options_data_list):
    logger.info(f"Preparing Excel Data for {len(options_data_list)} Option(s) on sheet '{sheet_name_target}'.")
    fixed_pm_quantity = 1000
    workbook = None
    all_data_for_pm_list_of_lists = []
    option_row_counts = [] 

    try:
        logger.debug("Loading workbook...")
        workbook = openpyxl.load_workbook(file_path)
        sheet_target = workbook[sheet_name_target]
        sheet_source_pnl = workbook[pnl_sheet_name_source]

        logger.debug(f"Reading 12 source values from '{pnl_sheet_name_source}' column '{PNL_BUY_SOURCE_COL}'...")
        pnl_buy_values = []
        if len(PNL_BUY_SOURCE_ROWS) != 12:
            logger.error(f"PNL_BUY_SOURCE_ROWS should have 12 rows, has {len(PNL_BUY_SOURCE_ROWS)}.")
            workbook.close(); return None, None
            
        for row_num_pnl in PNL_BUY_SOURCE_ROWS:
            cell_addr_pnl = f"{PNL_BUY_SOURCE_COL}{row_num_pnl}"
            value = sheet_source_pnl[cell_addr_pnl].value
            if isinstance(value, str):
                try: num_val = float(value); pnl_buy_values.append(num_val)
                except ValueError: pnl_buy_values.append(value) 
            elif value is not None: pnl_buy_values.append(value)
            else: pnl_buy_values.append('')
        
        for opt_data in options_data_list:
            option_idx = opt_data['id']
            cols = OPTION_COLUMN_MAP[option_idx]
            phase_config = PHASE_CONFIGS[opt_data['phase']]
            phase_dependent_block_start_row = phase_config['excel_start_row'] 
            
            logger.info(f"Processing Option {option_idx + 1} (Desc: '{opt_data['desc']}', Qty: {opt_data['qty']}, Phase: {opt_data['phase']}) for sheet '{sheet_name_target}'")

            user_qty_cell = f"{cols['user_qty_col']}2"
            sheet_target[user_qty_cell] = opt_data['qty']

            pm_qty_col_letter = cols['pm_qty_col']
            pm_desc_col_letter = cols['pm_desc_col']
            pm_third_col_letter = cols['pm_copy_end_col']

            for i in range(12): 
                current_row_for_qty_desc = phase_dependent_block_start_row + i
                if current_row_for_qty_desc <= EXCEL_PM_DATA_END_ROW:
                    sheet_target[f"{pm_qty_col_letter}{current_row_for_qty_desc}"] = fixed_pm_quantity
                    sheet_target[f"{pm_desc_col_letter}{current_row_for_qty_desc}"] = opt_data['desc']
                
                fixed_row_for_pnl_value = PNL_VALUE_TARGET_START_ROW_SHEET2 + i 
                sheet_target[f"{pm_third_col_letter}{fixed_row_for_pnl_value}"] = pnl_buy_values[i]

        logger.info(f"Reading data from '{sheet_name_target}' for PM consolidation...")
        for opt_data in options_data_list:
            option_idx = opt_data['id']
            cols = OPTION_COLUMN_MAP[option_idx]
            phase_config = PHASE_CONFIGS[opt_data['phase']]
            block_start_row_for_read = phase_config['excel_start_row'] 
            
            num_rows_this_option = EXCEL_PM_DATA_END_ROW - block_start_row_for_read + 1
            if num_rows_this_option < 0: num_rows_this_option = 0 
            option_row_counts.append(num_rows_this_option) 
            
            read_qty_col = cols['pm_qty_col']
            read_desc_col = cols['pm_desc_col']
            read_val_col = cols['pm_copy_end_col']

            option_rows_data = []
            for i in range(num_rows_this_option):
                current_row_to_read = block_start_row_for_read + i
                qty_val = sheet_target[f"{read_qty_col}{current_row_to_read}"].value
                desc_val = sheet_target[f"{read_desc_col}{current_row_to_read}"].value
                underlying_val = sheet_target[f"{read_val_col}{current_row_to_read}"].value 
                option_rows_data.append([qty_val if qty_val is not None else '', desc_val if desc_val is not None else '', underlying_val])
            all_data_for_pm_list_of_lists.extend(option_rows_data)
        
        if not all_data_for_pm_list_of_lists:
            logger.error(f"No data collected from '{sheet_name_target}' for PM DataFrame."); workbook.close(); return None, None
        
        consolidated_df = pd.DataFrame(all_data_for_pm_list_of_lists)
        
        if not consolidated_df.empty:
            consolidated_df.columns = ["Trade Amount", "Trade Description", "Underlying_Raw"]
        
        if not consolidated_df.empty and "Underlying_Raw" in consolidated_df.columns:
            def format_for_pm(x):
                if pd.isnull(x) or x == '': return '' 
                try: return f"{float(x):.3f}".replace('.', '-')
                except ValueError: return str(x).replace('.', '-')
            consolidated_df["Underlying"] = consolidated_df["Underlying_Raw"].apply(format_for_pm)
            consolidated_df.drop(columns=["Underlying_Raw"], inplace=True)
        
        pyperclip.copy(consolidated_df.to_csv(sep='\t', index=False, header=False))
        logger.info("Consolidated data copied to clipboard.")
        logger.info(f"Saving Excel workbook after setup on '{sheet_name_target}'...")
        workbook.save(file_path)
        logger.info("Excel workbook saved.")
        return consolidated_df, option_row_counts

    except Exception as e:
        logger.error(f"Error during Excel setup/consolidation for '{sheet_name_target}': {e}", exc_info=True); return None, None
    finally:
        if workbook: workbook.close()

def browser_operations_and_copy(url_to_open, total_rows_pasted_to_pm):
    # (Browser operations logic remains the same as previous version)
    try:
        logger.info(f"Opening URL: {url_to_open} ..."); webbrowser.open(url_to_open, new=2); time.sleep(WAIT_FOR_BROWSER_TO_OPEN)
        logger.debug("Navigating for paste: 8 TABs, 1 DOWN");
        for _ in range(8): send_keys('{TAB}', pause=KEY_PRESS_PAUSE); time.sleep(WAIT_AFTER_BROWSER_NAVIGATION)
        send_keys('{DOWN}', pause=KEY_PRESS_PAUSE); time.sleep(WAIT_AFTER_BROWSER_NAVIGATION)
        logger.debug("Pasting data (Ctrl+V)..."); send_keys('^v', pause=KEY_PRESS_PAUSE); time.sleep(WAIT_AFTER_BROWSER_PASTE); 
        logger.info("Paste complete.")
        logger.debug("Navigating for copy: 8 RIGHT arrows");
        for _ in range(8): send_keys('{RIGHT}', pause=KEY_PRESS_PAUSE); time.sleep(WAIT_AFTER_BROWSER_NAVIGATION)
        
        pm_down_arrow_count = total_rows_pasted_to_pm - 1 if total_rows_pasted_to_pm > 0 else 0
        logger.debug(f"Selecting results: SHIFT + {pm_down_arrow_count} DOWN, then SHIFT + 2 RIGHT")
        if pm_down_arrow_count > 0 : send_keys(f'+({{DOWN {pm_down_arrow_count}}})', pause=KEY_PRESS_PAUSE); time.sleep(WAIT_AFTER_BROWSER_NAVIGATION)
        send_keys(f'+({{RIGHT 2}})', pause=KEY_PRESS_PAUSE); time.sleep(WAIT_AFTER_BROWSER_NAVIGATION) 
        logger.debug("Copying results (Ctrl+C)..."); send_keys('^c', pause=KEY_PRESS_PAUSE); time.sleep(WAIT_FOR_COPY_OPERATION)

        pm_clipboard_content = pyperclip.paste()
        if pm_clipboard_content: pyperclip.copy(pm_clipboard_content); logger.info("PM results copied to clipboard.")
        else: logger.warning("Clipboard from PM was empty."); pyperclip.copy("")
        
        logger.debug("Closing browser tab (Ctrl+W)..."); send_keys('^w', pause=KEY_PRESS_PAUSE); time.sleep(WAIT_FOR_BROWSER_CLOSE); 
        logger.info("Browser close command sent.")
        return True
    except Exception as e: logger.error(f"Error during browser ops: {e}", exc_info=True); return False


def excel_distribute_pm_data_and_final_cleanup_openpyxl(file_path, sheet_name_target, all_options_inputs_list, pm_data_from_clipboard, option_row_counts_for_pm_results):
    # (This function remains largely the same as previous, focused on SHEET_NAME_PM_SETUP)
    workbook = None
    df_pm_results = None
    try:
        logger.info(f"Distributing PM Data to Excel sheet '{sheet_name_target}' & Final Cleanup.")
        if not pm_data_from_clipboard: logger.error("Clipboard empty (no PM data)."); return None

        logger.debug("Loading workbook for PM data distribution...")
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name_target]

        cleaned_lines = ['\t'.join(field.replace(',', '') for field in line.split('\t')) for line in pm_data_from_clipboard.splitlines()]
        cleaned_pm_data_for_df = "\n".join(cleaned_lines)
        
        df_pm_results = pd.read_csv(io.StringIO(cleaned_pm_data_for_df), sep='\t', header=None)
        logger.info(f"Parsed PM results from clipboard (Shape: {df_pm_results.shape})")

        if not df_pm_results.empty:
            df_pm_results.columns = ["DV01 Gamma", "Theta", "Vega"] 
            logger.debug(f"PM results DataFrame columns renamed. Head:\n{df_pm_results.head().to_string()}")

        current_row_index_in_pm_df = 0
        for idx, opt_data in enumerate(all_options_inputs_list):
            option_id_from_input = opt_data['id'] 
            phase_config = PHASE_CONFIGS[opt_data['phase']]
            num_rows_for_this_option_results = option_row_counts_for_pm_results[idx] 

            cols_map = OPTION_COLUMN_MAP[option_id_from_input] 
            target_paste_start_col_letter = cols_map['pm_result_start_col']
            target_result_end_col_letter = cols_map['pm_result_end_col'] 
            target_data_paste_start_row = phase_config['excel_results_data_start_row']
            
            logger.info(f"Processing PM results for Input Option {opt_data['id'] + 1} (Phase {opt_data['phase']}) on sheet '{sheet_name_target}':")

            if current_row_index_in_pm_df + num_rows_for_this_option_results > len(df_pm_results):
                logger.error(f"Not enough rows in PM results for Option {option_id_from_input + 1}."); 
                current_row_index_in_pm_df += num_rows_for_this_option_results 
                continue

            df_slice = df_pm_results.iloc[current_row_index_in_pm_df : current_row_index_in_pm_df + num_rows_for_this_option_results]
            if df_slice.empty: 
                logger.warning(f"Slice for Option {option_id_from_input + 1} is empty."); 
                current_row_index_in_pm_df += num_rows_for_this_option_results; continue
            
            start_col_idx_paste = column_index_from_string(target_paste_start_col_letter)
            for r_offset, pm_row_tuple in enumerate(df_slice.itertuples(index=False, name=None)):
                excel_row = target_data_paste_start_row + r_offset
                for c_offset, cell_val in enumerate(pm_row_tuple):
                    try: numeric_cell_val = float(cell_val); sheet.cell(row=excel_row, column=start_col_idx_paste + c_offset, value=numeric_cell_val)
                    except (ValueError, TypeError): sheet.cell(row=excel_row, column=start_col_idx_paste + c_offset, value=cell_val) 
            current_row_index_in_pm_df += num_rows_for_this_option_results

            if target_data_paste_start_row > 1: 
                cleanup_above_end_row = target_data_paste_start_row - 1
                clean_start_col_idx = column_index_from_string(target_paste_start_col_letter)
                clean_end_col_idx = column_index_from_string(target_result_end_col_letter)
                for r_val in range(1, cleanup_above_end_row + 1):
                    for c_idx in range(clean_start_col_idx, clean_end_col_idx + 1):
                        sheet.cell(row=r_val, column=c_idx).value = None
        
        num_options_processed = len(all_options_inputs_list) 
        for i in range(num_options_processed, 3):
            cols_map_unused = OPTION_COLUMN_MAP[i]
            unused_start = cols_map_unused['pm_result_start_col']; unused_end = cols_map_unused['pm_result_end_col']
            unused_start_idx = column_index_from_string(unused_start); unused_end_idx = column_index_from_string(unused_end)
            for r_val in range(1, EXCEL_PM_RESULTS_MAX_ROW_FOR_CLEARING_UNUSED + 1):
                for c_idx in range(unused_start_idx, unused_end_idx + 1): sheet.cell(row=r_val, column=c_idx).value = None
        
        logger.info(f"Saving Excel workbook after PM data distribution on '{sheet_name_target}'...")
        workbook.save(file_path)
        logger.info("Excel workbook saved.")
        return df_pm_results

    except Exception as e:
        logger.error(f"Error during Excel PM data distribution/cleanup for '{sheet_name_target}': {e}", exc_info=True); return None
    finally:
        if workbook: workbook.close()

def write_dataframes_to_sheet1(file_path: str, list_of_dataframes: list, target_sheet_name: str, start_row: int, row_offset: int):
    """
    Writes a list of DataFrames to a specified sheet in an Excel file, stacked vertically.
    Args:
        file_path (str): Path to the Excel file.
        list_of_dataframes (list): List of Pandas DataFrames to write.
        target_sheet_name (str): Name of the sheet to write to.
        start_row (int): Starting row for the first DataFrame.
        row_offset (int): Number of rows to offset for each subsequent DataFrame.
    """
    if not list_of_dataframes:
        logger.info("No DataFrames provided to write to Sheet1.")
        return

    logger.info(f"Writing {len(list_of_dataframes)} DataFrames to '{target_sheet_name}' in '{file_path}'.")
    workbook = None
    try:
        workbook = openpyxl.load_workbook(file_path)
        if target_sheet_name in workbook.sheetnames:
            sheet = workbook[target_sheet_name]
            # Optional: Clear previous content if needed. For now, we overwrite.
            # sheet.delete_rows(1, sheet.max_row) # Example: clear all rows
            logger.info(f"Using existing sheet: '{target_sheet_name}'.")
        else:
            sheet = workbook.create_sheet(target_sheet_name)
            logger.info(f"Created new sheet: '{target_sheet_name}'.")

        current_excel_row = start_row
        for i, df in enumerate(list_of_dataframes):
            if df.empty:
                logger.warning(f"DataFrame at index {i} is empty. Skipping write for this DataFrame.")
                # Optionally write a placeholder or just skip
                # For now, we just advance the row counter as if it took space, to maintain spacing
                # Or, we could choose not to advance if we want subsequent tables to be closer.
                # Let's assume we want to keep the spacing consistent.
                # current_excel_row += row_offset # This would create a gap even for an empty DF.
                # Let's skip advancing if DF is empty, so next DF uses this slot.
                continue 

            logger.info(f"Writing DataFrame {i+1} (Shape: {df.shape}) to '{target_sheet_name}' starting at row {current_excel_row}.")
            
            # Write headers and data
            # dataframe_to_rows writes header by default if index=False and header=True (which are defaults)
            rows = dataframe_to_rows(df, index=False, header=True)
            for r_idx, row in enumerate(rows, start=0): # r_idx is 0-based from dataframe_to_rows
                for c_idx, value in enumerate(row, start=1): # c_idx is 1-based for sheet.cell
                    sheet.cell(row=current_excel_row + r_idx, column=c_idx, value=value)
            
            # Advance the starting row for the next DataFrame
            # Add number of rows written (len(df) for data + 1 for header)
            # then add the fixed offset for spacing.
            # More simply, just use the fixed row_offset.
            current_excel_row += row_offset
            
        workbook.save(file_path)
        logger.info(f"Successfully wrote DataFrames to '{target_sheet_name}' and saved workbook.")

    except Exception as e:
        logger.error(f"Error writing DataFrames to '{target_sheet_name}': {e}", exc_info=True)
    finally:
        if workbook:
            workbook.close()


def run_pm_automation(options_data_from_dashboard: list):
    logger.info("Running Pricing Monkey Automation (Dashboard Integration)")
    if not options_data_from_dashboard: logger.warning("No option data from dashboard."); return None
    if not check_file_exists(FILE_PATH): return None

    consolidated_df_for_pm = None
    df_pm_results_full = None 
    option_row_counts_for_slicing = [] 

    try:
        consolidated_df_for_pm, option_row_counts_for_slicing = excel_setup_options_and_consolidate_for_pm_openpyxl(
            FILE_PATH, SHEET_NAME_PM_SETUP, PNL_SCENARIO_BUY_SHEET_NAME, options_data_from_dashboard
        )
        if consolidated_df_for_pm is None or consolidated_df_for_pm.empty or not option_row_counts_for_slicing:
            logger.error("Failed Excel data prep or no data/row_counts for PM."); return None
        
        total_rows_for_pm_selection = len(consolidated_df_for_pm)

        logger.info("Starting Browser Operations")
        browser_success = browser_operations_and_copy(PRICING_MONKEY_URL_FOR_PASTE, total_rows_for_pm_selection)
        if not browser_success: logger.error("Browser operations failed."); return None
        
        clipboard_content_from_pm = pyperclip.paste() 

        df_pm_results_full = excel_distribute_pm_data_and_final_cleanup_openpyxl(
            FILE_PATH, SHEET_NAME_PM_SETUP, options_data_from_dashboard, clipboard_content_from_pm, option_row_counts_for_slicing
        )

        if df_pm_results_full is None:
            logger.error("Failed to process PM data into Excel or get PM results DataFrame."); return None
        
        if len(consolidated_df_for_pm) != len(df_pm_results_full):
            logger.error(f"Row count mismatch! Input DF: {len(consolidated_df_for_pm)}, PM Results DF: {len(df_pm_results_full)}")
            return None

        consolidated_df_for_pm.reset_index(drop=True, inplace=True)
        df_pm_results_full.reset_index(drop=True, inplace=True)
        
        joined_df_full = pd.concat([consolidated_df_for_pm, df_pm_results_full], axis=1)
        logger.info(f"Successfully joined input data with PM results. Full joined shape: {joined_df_full.shape}")

        list_of_option_dfs = []
        current_pos = 0
        for count in option_row_counts_for_slicing:
            if count == 0: 
                empty_slice_df = pd.DataFrame(columns=joined_df_full.columns)
                list_of_option_dfs.append(empty_slice_df)
                logger.warning(f"Option contributed 0 rows. Appending empty DF slice.")
                continue

            option_df_slice = joined_df_full.iloc[current_pos : current_pos + count].copy()
            list_of_option_dfs.append(option_df_slice)
            current_pos += count
        
        if current_pos != len(joined_df_full) and sum(option_row_counts_for_slicing) != 0 : 
             logger.error(f"Mismatch after slicing. Expected {len(joined_df_full)}, processed {current_pos} from {option_row_counts_for_slicing}")
             return None

        logger.info(f"Successfully sliced joined_df into {len(list_of_option_dfs)} DataFrames for each option.")
        
        # --- NEW: Write the list of DataFrames to Sheet1 ---
        if list_of_option_dfs:
            write_dataframes_to_sheet1(FILE_PATH, list_of_option_dfs, TARGET_SHEET_FOR_OUTPUT, START_ROW_SHEET1, ROW_OFFSET_PER_DF_SHEET1)
        else:
            logger.warning("No option DataFrames to write to Sheet1.")
        # --- END NEW ---

        return list_of_option_dfs

    except Exception as e:
        logger.error(f"CRITICAL ERROR in run_pm_automation: {e}", exc_info=True); return None
    finally:
        logger.info("Pricing Monkey Automation Run Finished")


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(name)s - %(funcName)s - %(message)s')
    logger.info("Running pMoneyAuto.py in STANDALONE mode for testing")
    
    # (Standalone test input function remains the same as in pMoneyAuto_sliced_output)
    def get_all_user_inputs_for_standalone():
        num_options = int(input("How many options for standalone test (1-3)? ").strip())
        collected_options_data = []
        min_phase = 1
        for i in range(num_options):
            desc = input(f"Opt {i+1} Desc (e.g., '1st 10y note 0 out call'): ").strip()
            qty = int(input(f"Opt {i+1} Qty: ").strip())
            phase_str = input(f"Opt {i+1} Phase ({', '.join(map(str, PHASE_CONFIGS.keys()))}, >= {min_phase}): ").strip()
            while not phase_str.isdigit() or int(phase_str) not in PHASE_CONFIGS or int(phase_str) < min_phase:
                phase_str = input(f"Invalid. Phase ({', '.join(map(str, PHASE_CONFIGS.keys()))}, >= {min_phase}): ").strip()
            phase = int(phase_str)
            min_phase = phase
            collected_options_data.append({'desc': desc, 'qty': qty, 'phase': phase, 'id': i})
        return collected_options_data

    standalone_options_input = get_all_user_inputs_for_standalone()
    
    if standalone_options_input:
        list_of_final_dataframes = run_pm_automation(standalone_options_input)

        if list_of_final_dataframes:
            logger.info(f"\n=== FINAL DATAFRAMES (from standalone run) - Count: {len(list_of_final_dataframes)} ===")
            for i, df_opt in enumerate(list_of_final_dataframes):
                logger.info(f"\n--- Option {i+1} DataFrame (Shape: {df_opt.shape}) ---")
                logger.info(df_opt.to_string())
            logger.info("Check Sheet1 in the Excel file for the output.")
            logger.info("=====================================================")
        else:
            logger.error("Standalone run completed, but no final list of DataFrames was produced.")
    else:
        logger.warning("No inputs provided for standalone run. Exiting.")
    
    logger.info("Standalone pMoneyAuto.py test finished")
    time.sleep(1)
