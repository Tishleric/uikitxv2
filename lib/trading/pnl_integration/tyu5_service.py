"""TYU5 Service - Orchestrates P&L calculations

This service monitors data stores and triggers TYU5 calculations
when new data arrives.
"""

import logging
from datetime import datetime, date
from typing import Optional, Callable
from pathlib import Path
import threading
import time
import pandas as pd
import openpyxl
from openpyxl import load_workbook

from .tyu5_adapter import TYU5Adapter
from .bachelier_attribution import BachelierAttributionService
from .tyu5_database_writer import TYU5DatabaseWriter

logger = logging.getLogger(__name__)


class TYU5Service:
    """Service to manage TYU5 P&L calculations."""
    
    def __init__(self, 
                 db_path: Optional[str] = None,
                 output_dir: Optional[str] = None,
                 enable_attribution: bool = True,
                 enable_db_writer: bool = True):
        """Initialize the service.
        
        Args:
            db_path: Path to SQLite database
            output_dir: Directory for output files (defaults to data/output/pnl)
            enable_attribution: Enable Bachelier P&L attribution (for easy reversion)
            enable_db_writer: Enable database persistence of TYU5 results
        """
        self.adapter = TYU5Adapter(db_path)
        self.output_dir = Path(output_dir or "data/output/pnl")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Attribution service (can be disabled via flag)
        self.attribution_service = BachelierAttributionService(enable_attribution)
        self.enable_attribution = enable_attribution
        
        # Database writer service
        self.db_writer = TYU5DatabaseWriter(db_path) if enable_db_writer else None
        self.enable_db_writer = enable_db_writer
        
        # Monitoring state
        self._monitoring = False
        self._monitor_thread = None
        self._last_calculation_time = None
        
    def calculate_pnl(self, 
                     trade_date: Optional[date] = None,
                     output_format: str = "excel") -> str:
        """Run P&L calculation for a specific date.
        
        Args:
            trade_date: Date to calculate (None for all)
            output_format: Output format ("excel" or "csv")
            
        Returns:
            Path to output file
        """
        # Generate output filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        date_str = trade_date.strftime("%Y%m%d") if trade_date else "all"
        
        if output_format == "excel":
            output_file = self.output_dir / f"tyu5_pnl_{date_str}_{timestamp}.xlsx"
        else:
            output_file = self.output_dir / f"tyu5_pnl_{date_str}_{timestamp}.csv"
            
        # Run calculation
        logger.info(f"Starting TYU5 calculation for {trade_date or 'all dates'}")
        
        success = self.adapter.run_calculation(
            output_file=str(output_file),
            trade_date=trade_date
        )
        
        if success:
            self._last_calculation_time = datetime.now()
            logger.info(f"Calculation complete. Output: {output_file}")
            
            # Enhance with Bachelier attribution if enabled
            if self.enable_attribution and str(output_file).endswith('.xlsx'):
                try:
                    self._enhance_excel_with_attribution(str(output_file))
                except Exception as e:
                    logger.error(f"Failed to add attribution: {e}")
                    # Continue even if attribution fails
                    
            # Persist to database if enabled
            if self.enable_db_writer and self.db_writer and str(output_file).endswith('.xlsx'):
                try:
                    self._persist_to_database(str(output_file), datetime.now())
                except Exception as e:
                    logger.error(f"Failed to persist to database: {e}")
                    # Continue even if database write fails
                    
            return str(output_file)
        else:
            logger.error("Calculation failed")
            return ""
            
    def print_summary(self, trade_date: Optional[date] = None):
        """Print calculation summary to console.
        
        Args:
            trade_date: Date to summarize (None for all)
        """
        # Get data
        excel_data = self.adapter.prepare_excel_data(trade_date)
        summary = self.adapter.get_calculation_summary(excel_data)
        
        # Print summary
        print("\n" + "="*60)
        print("TYU5 P&L CALCULATION SUMMARY")
        print("="*60)
        
        if 'error' in summary:
            print(f"ERROR: {summary['error']}")
            return
            
        print(f"Total Trades: {summary['total_trades']}")
        print(f"Unique Symbols: {summary['unique_symbols']}")
        print(f"Date Range: {summary['date_range']['start']} to {summary['date_range']['end']}")
        print(f"  - Buys: {summary['total_buys']}")
        print(f"  - Sells: {summary['total_sells']}")
        print(f"  - Futures: {summary['futures_count']}")
        print(f"  - Options: {summary['options_count']}")
        print(f"Market Prices Loaded: {summary['prices_loaded']}")
        print("="*60)
        
    def start_monitoring(self, 
                        check_interval: int = 60,
                        callback: Optional[Callable] = None):
        """Start monitoring for new data.
        
        Args:
            check_interval: Seconds between checks
            callback: Optional callback when calculation completes
        """
        if self._monitoring:
            logger.warning("Monitoring already active")
            return
            
        self._monitoring = True
        self._monitor_thread = threading.Thread(
            target=self._monitor_loop,
            args=(check_interval, callback),
            daemon=True
        )
        self._monitor_thread.start()
        logger.info(f"Started monitoring with {check_interval}s interval")
        
    def stop_monitoring(self):
        """Stop monitoring for new data."""
        self._monitoring = False
        if self._monitor_thread:
            self._monitor_thread.join(timeout=5)
        logger.info("Stopped monitoring")
        
    def _monitor_loop(self, interval: int, callback: Optional[Callable]):
        """Monitor loop that checks for new data.
        
        Args:
            interval: Seconds between checks
            callback: Optional callback function
        """
        last_check = {}
        
        while self._monitoring:
            try:
                # Check for new data (simplified for now)
                # In production, this would check row counts or timestamps
                
                # For now, just run calculation if we haven't in a while
                if self._last_calculation_time is None or \
                   (datetime.now() - self._last_calculation_time).seconds > interval:
                    
                    logger.info("Running scheduled calculation")
                    output_file = self.calculate_pnl()
                    
                    if output_file and callback:
                        callback(output_file)
                        
            except Exception as e:
                logger.error(f"Error in monitor loop: {e}")
                
            time.sleep(interval)
            
    def test_connection(self) -> bool:
        """Test database connection and data availability.
        
        Returns:
            True if connection successful and data available
        """
        try:
            # Try to get some data
            excel_data = self.adapter.prepare_excel_data()
            
            has_trades = not excel_data['Trades_Input'].empty
            has_prices = not excel_data['Market_Prices'].empty
            
            logger.info(f"Connection test: Trades={has_trades}, Prices={has_prices}")
            
            return has_trades or has_prices
            
        except Exception as e:
            logger.error(f"Connection test failed: {e}")
            return False
            
    def _enhance_excel_with_attribution(self, excel_file: str):
        """Enhance Excel file with Bachelier P&L attribution.
        
        Args:
            excel_file: Path to the Excel file to enhance
        """
        logger.info(f"Enhancing Excel with Bachelier attribution: {excel_file}")
        
        # Read existing Excel file
        xl_file = pd.ExcelFile(excel_file)
        
        # Check if Positions sheet exists
        if 'Positions' not in xl_file.sheet_names:
            logger.warning("No Positions sheet found, skipping attribution")
            return
            
        # Read positions 
        positions_df = pd.read_excel(excel_file, sheet_name='Positions')
        
        # Get market prices from the adapter (database)
        market_prices_df = self.adapter.get_market_prices()
        
        if market_prices_df is None or market_prices_df.empty:
            logger.warning("No market prices available in database, skipping attribution")
            return
            
        # Enhance positions with attribution
        enhanced_positions = self.attribution_service.enhance_positions_dataframe(
            positions_df, market_prices_df
        )
        
        # Write back to Excel
        # Use openpyxl to preserve other sheets
        wb = load_workbook(excel_file)
        
        # Remove old Positions sheet
        if 'Positions' in wb.sheetnames:
            del wb['Positions']
            
        # Create new Positions sheet with attribution
        ws = wb.create_sheet('Positions', 2)  # Insert at position 2
        
        # Write headers
        headers = list(enhanced_positions.columns)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
            
        # Write data
        for row_idx, row in enhanced_positions.iterrows():
            for col_idx, value in enumerate(row, 1):
                # Handle NaN values
                if pd.isna(value):
                    ws.cell(row=row_idx+2, column=col_idx, value='')
                else:
                    ws.cell(row=row_idx+2, column=col_idx, value=value)
                    
        # Format attribution columns
        attribution_cols = ['delta_pnl', 'gamma_pnl', 'vega_pnl', 'theta_pnl', 
                           'speed_pnl', 'residual', 'total_attributed_pnl']
        
        # Find column indices
        col_indices = {}
        for col in attribution_cols:
            if col in headers:
                col_indices[col] = headers.index(col) + 1
                
        # Apply number formatting to attribution columns
        for col_name, col_idx in col_indices.items():
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.number_format = '#,##0.00'
                
        # Add conditional formatting for attribution_calculated column
        if 'attribution_calculated' in headers:
            calc_col_idx = headers.index('attribution_calculated') + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=calc_col_idx)
                if cell.value is True:
                    cell.value = 'Yes'
                elif cell.value is False:
                    cell.value = 'No'
                    
        # Auto-fit columns (approximate)
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Save the enhanced Excel file
        wb.save(excel_file)
        wb.close()
        
        logger.info(f"Successfully enhanced Excel with attribution for {len(enhanced_positions)} positions") 

    def _persist_to_database(self, excel_file: str, calc_timestamp: datetime):
        """Read TYU5 Excel output and persist to database.
        
        Args:
            excel_file: Path to TYU5 Excel output
            calc_timestamp: Timestamp of the calculation
        """
        try:
            # Read all sheets from Excel
            excel_data = pd.read_excel(excel_file, sheet_name=None)
            
            # Extract dataframes
            positions_df = excel_data.get('Positions', pd.DataFrame())
            trades_df = excel_data.get('Trades', pd.DataFrame())
            breakdown_df = excel_data.get('Position_Breakdown', pd.DataFrame())
            risk_df = excel_data.get('Risk_Matrix', pd.DataFrame())
            summary_df = excel_data.get('Summary', pd.DataFrame())
            
            # Write to database
            success = self.db_writer.write_results(
                positions_df=positions_df,
                trades_df=trades_df,
                breakdown_df=breakdown_df,
                risk_df=risk_df,
                summary_df=summary_df,
                calc_timestamp=calc_timestamp
            )
            
            if success:
                logger.info(f"Successfully persisted TYU5 results to database")
            else:
                logger.error("Failed to persist TYU5 results to database")
                
        except Exception as e:
            logger.error(f"Error reading Excel for database persistence: {e}")
            raise 