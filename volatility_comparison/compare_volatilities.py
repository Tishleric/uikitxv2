"""
Calculate theoretical volatility and compare with Actant and PM
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
from bond_option_pricer import calculate_bond_option_volatility, parse_treasury_price

def create_pivot_table(comparison_df):
    """
    Transform comparison data from row-per-option to column-per-option format.
    
    Args:
        comparison_df: DataFrame with comparison results
        
    Returns:
        DataFrame in pivot format with options as columns and metrics as rows
    """
    # Separate Actant and PM data
    actant_data = comparison_df[comparison_df['Source'] == 'Actant'].copy()
    pm_data = comparison_df[comparison_df['Source'] == 'PM'].copy()
    
    # Create unified data structure
    all_options = []
    
    # Process Actant data
    for _, row in actant_data.iterrows():
        option_data = {
            'Option_ID': row['Description'],  # XCME.ZN, 13JUN25, etc.
            'Source': 'Actant',
            'Description': row['Description'],
            'Expiry_Date': row['Description'],  # Keep scenario name as expiry for Actant
            'Strike': row['Strike'],
            'Future_Price': row['Future_Price'],
            'Days': row['Days'],
            'Time_Years': row['Time_Years'],
            'Market_Price': row['Market_Price'],
            'Market_Price_64ths': row['Market_Price_64ths'],
            'Bid': row['Bid'],
            'Ask': row['Ask'],
            'Market_Vol': row['Market_Vol_%'],
            'Calculated_Vol': row['Calculated_Vol_%'],
            'PM_Market_Vol': np.nan  # Will be filled later if matching PM option exists
        }
        all_options.append(option_data)
    
    # Process PM data
    for _, row in pm_data.iterrows():

        
        option_data = {
            'Option_ID': row['Description'],  # 1st 10y note 25 out call, etc.
            'Source': 'PM',
            'Description': row['Description'],
            'Expiry_Date': row.get('Expiry_Date', ''),  # Use actual expiry column from PM data
            'Strike': row['Strike'],
            'Future_Price': row['Future_Price'],
            'Days': row['Days'],
            'Time_Years': row['Time_Years'],
            'Market_Price': row['Market_Price'],
            'Market_Price_64ths': row['Market_Price_64ths'],
            'Bid': row['Bid'],
            'Ask': row['Ask'],
            'Market_Vol': row['Market_Vol_%'],
            'Calculated_Vol': row['Calculated_Vol_%'],
            'PM_Market_Vol': row['Market_Vol_%']  # PM options reference themselves
        }
        all_options.append(option_data)
    
    # Create pivot DataFrame
    options_df = pd.DataFrame(all_options)
    
    # Transpose to get options as columns
    pivot_data = {}
    
    # Add each option as a column
    for _, option in options_df.iterrows():
        option_id = option['Option_ID']
        
        # Create column data
        column_data = {
            'Source': option['Source'],
            'Expiry_Date': option['Expiry_Date'],
            'Strike': option['Strike'],
            'Future_Price': option['Future_Price'],
            'Days': option['Days'],
            'Time_Years': option['Time_Years'],
            'Market_Price': option['Market_Price'],
            'Market_Price_64ths': option['Market_Price_64ths'],
            'Bid': option['Bid'],
            'Ask': option['Ask'],
            'Market_Vol': option['Market_Vol'],
            'Calculated_Vol': option['Calculated_Vol']
        }
        
        pivot_data[option_id] = column_data
    
    # Convert to DataFrame with proper structure
    pivot_df = pd.DataFrame(pivot_data)
    
    # Calculate difference rows
    if not pivot_df.empty:
        # Calculate "Difference from Calculated" row
        diff_from_calc = {}
        
        for col in pivot_df.columns:
            option_data = pivot_df[col]
            market_vol = option_data['Market_Vol']
            calc_vol = option_data['Calculated_Vol']
            
            # Difference from Calculated = Market_Vol - Calculated_Vol
            if pd.notna(market_vol) and pd.notna(calc_vol) and calc_vol != '':
                try:
                    market_val = float(market_vol)
                    calc_val = float(calc_vol)
                    diff_from_calc[col] = market_val - calc_val
                except:
                    diff_from_calc[col] = ''
            else:
                diff_from_calc[col] = ''
        
        # Add difference row to pivot DataFrame
        pivot_df.loc['Difference_from_Calculated'] = pd.Series(diff_from_calc)
    
    return pivot_df

def main():
    # Capture run timestamp
    run_timestamp = datetime.now()
    
    # Load data
    actant_df = pd.read_csv('actant_data.csv') if os.path.exists('actant_data.csv') else pd.DataFrame()
    pm_df = pd.read_csv('pm_data.csv') if os.path.exists('pm_data.csv') else pd.DataFrame()
    
    # Read timestamps
    actant_timestamp = None
    pm_timestamp = None
    
    if os.path.exists('actant_timestamp.txt'):
        with open('actant_timestamp.txt', 'r') as f:
            actant_timestamp = f.read().strip()
    
    if os.path.exists('pm_timestamp.txt'):
        with open('pm_timestamp.txt', 'r') as f:
            pm_timestamp = f.read().strip()
    
    if actant_df.empty:
        print("ERROR: No Actant data found in actant_data.csv")
        return
    
    print(f"Loaded {len(actant_df)} Actant records")
    if actant_timestamp:
        print(f"Actant data captured at: {actant_timestamp}")
    
    print(f"Loaded {len(pm_df)} PM records")
    if pm_timestamp:
        print(f"PM data captured at: {pm_timestamp}")
    
    # Create comparison results
    comparison_results = []
    
    # Process Actant data
    for _, actant_row in actant_df.iterrows():
        comparison_results.append({
            'Source': 'Actant',
            'Description': actant_row.get('scenario', 'Unknown'),
            'Expiry_Date': actant_row.get('scenario', 'Unknown'),  # Use scenario name as expiry for Actant
            'Strike': actant_row['K'],
            'Future_Price': actant_row['F'],
            'Days': actant_row['T'] * 365,
            'Time_Years': actant_row['T'],
            'Market_Price': '',
            'Market_Price_64ths': '',
            'Bid': '',
            'Ask': '',
            'Market_Vol_%': actant_row['Vol'],
            'Calculated_Vol_%': '',
            'Vol_Abs_Diff_%': ''
        })
    
    # Process PM data with theoretical calculations
    print("\n" + "="*80)
    print("PROCESSING PRICING MONKEY DATA - CALCULATED VOLATILITY CALCULATIONS")
    print("="*80)
    
    for i, pm_row in pm_df.iterrows():
        print(f"\n--- PM Option {i+1}: {pm_row['description']} ---")
        
        # Use the parsed underlying_decimal from PM scraper
        underlying_price = pm_row['underlying_decimal']
        print(f"Future Price (F): {underlying_price:.6f}")
        
        # Strike is already parsed as float in PM scraper
        strike = pm_row['strike'] 
        print(f"Strike Price (K): {strike:.6f}")
        
        # Convert days to years for time parameter
        days = float(pm_row['days']) if pm_row['days'] else 0
        T = days / 252.0  # Convert business days to years
        print(f"Days to Expiry: {days}")
        print(f"Time to Expiry (T): {T:.6f} years")
        
        # Parse bid/ask prices (they're strings like "05.5", "09", etc.)
        def parse_price_string(price_str):
            """Convert price strings like '05.5' or '09' to float"""
            try:
                if isinstance(price_str, str):
                    return float(price_str)
                return float(price_str) if not pd.isna(price_str) else 0
            except:
                return 0
        
        bid = parse_price_string(pm_row['bid'])
        ask = parse_price_string(pm_row['ask'])
        price = parse_price_string(pm_row['price'])
        
        # Calculate midpoint (use price if bid/ask unavailable)
        if bid > 0 and ask > 0:
            mid_price = (bid + ask) / 2.0
        else:
            mid_price = price
        
        print(f"Price: {price}, Bid: {bid}, Ask: {ask}")
        print(f"Market Price (midpoint): {mid_price:.6f} (in 64ths format)")
        print(f"Market Vol: {pm_row['vol']:.2f}%")
        
        # Calculate calculated volatility using our bond option pricer
        calc_vol = np.nan
        error_msg = ""
        
        if T > 0 and mid_price > 0 and underlying_price > 0 and strike > 0:
            try:
                print(f"\nCalling calculate_bond_option_volatility with:")
                print(f"  F = {underlying_price:.6f}")
                print(f"  K = {strike:.6f}")
                print(f"  T = {T:.6f}")
                print(f"  market_price = {mid_price:.6f}")
                print(f"  option_type = 'call'")
                
                calc_vol = calculate_bond_option_volatility(
                    F=underlying_price,
                    K=strike,
                    T=T,  # Time in years
                    market_price=mid_price,  # Market price in 64ths
                    option_type='call'
                )
                print(f"  ✅ RESULT: Calculated Vol = {calc_vol:.6f}")
                
            except Exception as e:
                error_msg = str(e)
                print(f"  ❌ ERROR: {error_msg}")
        else:
            missing = []
            if T <= 0: missing.append("T <= 0")
            if mid_price <= 0: missing.append("mid_price <= 0") 
            if underlying_price <= 0: missing.append("underlying_price <= 0")
            if strike <= 0: missing.append("strike <= 0")
            error_msg = f"Invalid inputs: {', '.join(missing)}"
            print(f"  ❌ SKIPPED: {error_msg}")
        
        # Calculate absolute difference
        vol_diff = abs(pm_row['vol'] - calc_vol) if not np.isnan(calc_vol) else np.nan
        
        # Format calculated vol output
        calc_vol_str = f"{calc_vol:.2f}" if not np.isnan(calc_vol) else "N/A"
        vol_diff_str = f"{vol_diff:.2f}" if not np.isnan(vol_diff) else "N/A"
        
        print(f"Market Vol: {pm_row['vol']:.2f}% vs Calculated Vol: {calc_vol_str}%")
        print(f"Absolute Difference: {vol_diff_str}%")
        
        
        comparison_results.append({
            'Source': 'PM',
            'Description': pm_row['description'],
            'Expiry_Date': pm_row['expiry'],  # Add expiry date from PM data
            'Strike': strike,
            'Future_Price': underlying_price,
            'Days': days,
            'Time_Years': T,
            'Market_Price': mid_price,
            'Market_Price_64ths': f"{mid_price:.2f}",
            'Bid': bid,
            'Ask': ask,
            'Market_Vol_%': pm_row['vol'],
            'Calculated_Vol_%': f"{calc_vol:.2f}" if not np.isnan(calc_vol) else error_msg,
            'Vol_Abs_Diff_%': f"{vol_diff:.2f}" if not np.isnan(vol_diff) else ''
        })
    
    # Create DataFrame and save to Excel
    comparison_df = pd.DataFrame(comparison_results)
    
    # Create pivot table format
    pivot_df = create_pivot_table(comparison_df)
    
    # Generate timestamped filename
    filename_timestamp = run_timestamp.strftime('%Y%m%d_%H%M%S')
    excel_filename = f'volatility_comparison_{filename_timestamp}.xlsx'
    
    # Save to Excel with enhanced formatting
    with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
        # Save pivot format
        if not pivot_df.empty:
            # Add timestamp rows at the top of the pivot table
            timestamp_data = {}
            for col in pivot_df.columns:
                timestamp_data[col] = ''
            
            # Create timestamp rows
            actant_time_row = pd.Series(timestamp_data)
            pm_time_row = pd.Series(timestamp_data)
            
            # Set first column values
            actant_time_row.iloc[0] = f'Actant Timestamp: {actant_timestamp}' if actant_timestamp else 'Actant Timestamp: N/A'
            pm_time_row.iloc[0] = f'PM Timestamp: {pm_timestamp}' if pm_timestamp else 'PM Timestamp: N/A'
            
            # Insert timestamp rows at the beginning
            pivot_with_timestamps = pd.DataFrame()
            pivot_with_timestamps = pd.concat([
                pd.DataFrame({'Actant_Timestamp': actant_time_row}).T,
                pd.DataFrame({'PM_Timestamp': pm_time_row}).T,
                pd.DataFrame({'': pd.Series(timestamp_data)}).T,  # Empty row for spacing
                pivot_df
            ])
            
            pivot_with_timestamps.to_excel(writer, sheet_name='Volatility_Comparison', index=True)
            
            # Apply comprehensive formatting
            from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
            
            worksheet = writer.sheets['Volatility_Comparison']
            
            # Define styles
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            right_align = Alignment(horizontal='right', vertical='center')
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Highlight colors for key rows
            market_vol_fill = PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid')  # Light blue
            calc_vol_fill = PatternFill(start_color='E8F6F3', end_color='E8F6F3', fill_type='solid')   # Light green
            
            # Get data dimensions
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # Apply borders to all cells
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.border = thin_border
            
            # Special formatting for timestamp rows (rows 2 and 3)
            for row in [2, 3]:  # Timestamp rows
                for col in range(1, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    if col == 1:  # First column with timestamp text
                        cell.font = Font(bold=True, italic=True)
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Format headers (now at row 5 after timestamps and empty row)
            header_row = 5
            for col in range(1, max_col + 1):
                cell = worksheet.cell(row=header_row, column=col)
                cell.alignment = center_align
                cell.font = Font(bold=True)
            
            # Format index column (column A) - left align, bold
            for row in range(header_row + 1, max_row + 1):
                cell = worksheet.cell(row=row, column=1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Right-align all data cells (columns B onwards, rows after header)
            for row in range(header_row + 1, max_row + 1):
                for col in range(2, max_col + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = right_align
            
            # Highlight Market_Vol and Calculated_Vol rows
            for row in range(header_row + 1, max_row + 1):
                row_label = worksheet.cell(row=row, column=1).value
                if row_label == 'Market_Vol':
                    for col in range(1, max_col + 1):
                        worksheet.cell(row=row, column=col).fill = market_vol_fill
                elif row_label == 'Calculated_Vol':
                    for col in range(1, max_col + 1):
                        worksheet.cell(row=row, column=col).fill = calc_vol_fill
            
            # Auto-fit columns
            for col_num, column in enumerate(pivot_df.columns, 2):  # Start from column B (2)
                # Convert column number to Excel column letter
                if col_num <= 26:
                    col_letter = chr(64 + col_num)
                else:
                    # For columns beyond Z (AA, AB, etc.)
                    first_letter = chr(64 + ((col_num - 1) // 26))
                    second_letter = chr(64 + ((col_num - 1) % 26) + 1)
                    col_letter = first_letter + second_letter
                
                # Calculate max width needed
                max_width = len(str(column)) + 2
                for row_data in pivot_df[column]:
                    if pd.notna(row_data):
                        max_width = max(max_width, len(str(row_data)) + 2)
                
                worksheet.column_dimensions[col_letter].width = min(max_width, 35)
            
            # Set index column width - wider to accommodate timestamps
            worksheet.column_dimensions['A'].width = 30
            
        else:
            # Fallback to original format if pivot fails
            comparison_df.to_excel(writer, sheet_name='Volatility_Comparison', index=False)
    
    print(f"\nCreated {excel_filename} with pivot table format and timestamps")
    
    # Also save a copy with the fixed name for compatibility
    try:
        pivot_df.to_excel('volatility_comparison_formatted.xlsx', sheet_name='Volatility_Comparison', index=True)
        print("Also saved as volatility_comparison_formatted.xlsx for compatibility")
    except PermissionError:
        print("Note: Could not save volatility_comparison_formatted.xlsx (file may be open)")
    
    # Print summary
    print("\nVOLATILITY COMPARISON SUMMARY")
    print("=" * 50)
    print(f"Actant options: {len(actant_df)}")
    print(f"PM options: {len(pm_df)}")
    
    # Print PM calculated calculations
    pm_data = comparison_df[comparison_df['Source'] == 'PM']
    if not pm_data.empty:
        print(f"\nCalculated Calculations (using UIKitXv2 analyze_bond_future_option_greeks):")
        for _, row in pm_data.iterrows():
            if row['Calculated_Vol_%']:
                print(f"\n{row['Description']}:")
                print(f"  Strike: {row['Strike']}")
                print(f"  Underlying: {row['Future_Price']:.4f}")
                print(f"  Mid Price: {row['Market_Price']:.6f} ({row['Market_Price_64ths']}/64ths)")
                print(f"  Market Vol: {row['Market_Vol_%']:.2f}%")
                print(f"  Calculated Vol: {row['Calculated_Vol_%']}%")
                print(f"  Absolute Difference: {row['Vol_Abs_Diff_%']}%")

if __name__ == "__main__":
    main() 