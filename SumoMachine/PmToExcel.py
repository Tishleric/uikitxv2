"""
PmToExcel - Standalone Pricing Monkey to Excel Automation

This script automates the process of extracting data from Pricing Monkey
and exporting it to an Excel file. It performs browser automation to
copy data and processes it with specific transformations for futures pricing.
"""

import time
import io
import os
import sys
import pandas as pd
import pyperclip
import webbrowser
import traceback
import logging
from datetime import datetime
from pywinauto.keyboard import send_keys
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import numbers

# --- Configuration ---
PRICING_MONKEY_URL = "https://pricingmonkey.com/b/81a1616a-0ccc-464a-a3b6-2138be32aceb"
EXCEL_FILE_PATH = r"C:\Users\g.shah\Documents\Pm2Excel\ActantBackend.xlsx"
TARGET_SHEET = "Sheet1"
TARGET_CELL = "A2"

# Expected columns from Pricing Monkey
EXPECTED_COLUMNS = ["Notes", "Trade Amount", "Price", "DV01", "DV01 Gamma"]

# Timing constants (in seconds)
WAIT_FOR_BROWSER_TO_OPEN = 3.0
WAIT_AFTER_NAVIGATION = 0.1
WAIT_AFTER_PASTE = 0.2
WAIT_FOR_COPY_OPERATION = 0.5
WAIT_FOR_BROWSER_CLOSE = 0.5
KEY_PRESS_PAUSE = 0.01

# Configure logging
LOG_DIR = r"C:\Users\g.shah\Documents\Pm2Excel"
# Create log directory if it doesn't exist
os.makedirs(LOG_DIR, exist_ok=True)
log_filename = os.path.join(LOG_DIR, f"PmToExcel_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(funcName)s - %(message)s',
    handlers=[
        logging.FileHandler(log_filename),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


def parse_futures_price(price_str):
    """
    Convert price format to decimal.
    
    Handles format like "123-456" where the calculation is XXX + YYY/32.
    The number of digits after the hyphen is variable.
    
    Args:
        price_str: String representation of the price
        
    Returns:
        float: Decimal representation of the price
    """
    if pd.isna(price_str) or price_str == "":
        return 0.0
        
    price_str = str(price_str).strip()
    
    # Check if it contains a hyphen
    if '-' in price_str:
        try:
            parts = price_str.split('-')
            xxx_part = float(parts[0])
            
            # Handle the YYY part
            if len(parts) > 1 and parts[1]:
                yyy_part = float(parts[1])
                decimal_value = xxx_part + (yyy_part / 32.0)
                return decimal_value
            else:
                return xxx_part
                
        except (ValueError, IndexError) as e:
            logger.warning(f"Could not parse price '{price_str}': {e}")
            return 0.0
    else:
        # Regular decimal format
        try:
            return float(price_str)
        except ValueError:
            logger.warning(f"Could not parse price '{price_str}'")
            return 0.0


def clean_numeric_value(value):
    """
    Clean numeric values by removing commas and converting to float.
    
    Args:
        value: Value to clean (string or numeric)
        
    Returns:
        float: Cleaned numeric value
    """
    if pd.isna(value) or value == "":
        return 0.0
        
    # Convert to string and remove commas
    value_str = str(value).replace(',', '').strip()
    
    try:
        return float(value_str)
    except ValueError:
        logger.warning(f"Could not convert '{value}' to numeric")
        return 0.0


def browser_automation():
    """
    Automate browser interaction to copy data from Pricing Monkey.
    
    Returns:
        bool: Success status
    """
    try:
        logger.info(f"Opening URL: {PRICING_MONKEY_URL}")
        webbrowser.open(PRICING_MONKEY_URL, new=2)
        time.sleep(WAIT_FOR_BROWSER_TO_OPEN)
        
        # Navigate to the data area
        logger.info("Navigating with TAB (8 times)...")
        for i in range(8):
            send_keys('{TAB}', pause=KEY_PRESS_PAUSE)
            time.sleep(WAIT_AFTER_NAVIGATION)
        
        logger.info("Pressing DOWN arrow once...")
        send_keys('{DOWN}', pause=KEY_PRESS_PAUSE)
        time.sleep(WAIT_AFTER_NAVIGATION)
        
        # Select the data range
        logger.info("Selecting data: SHIFT + DOWN (11 times)...")
        send_keys('+({DOWN 11})', pause=KEY_PRESS_PAUSE)
        time.sleep(WAIT_AFTER_NAVIGATION)
        
        logger.info("Extending selection: SHIFT + RIGHT (4 times to select 5 columns total)...")
        send_keys('+({RIGHT 4})', pause=KEY_PRESS_PAUSE)
        time.sleep(WAIT_AFTER_NAVIGATION)
        
        # Copy the data
        logger.info("Copying data (Ctrl+C)...")
        send_keys('^c', pause=KEY_PRESS_PAUSE)
        time.sleep(WAIT_FOR_COPY_OPERATION)
        
        # Close the browser tab
        logger.info("Closing browser tab (Ctrl+W)...")
        send_keys('^w', pause=KEY_PRESS_PAUSE)
        time.sleep(WAIT_FOR_BROWSER_CLOSE)
        
        logger.info("Browser automation completed successfully")
        return True
        
    except Exception as e:
        logger.error(f"Error during browser automation: {e}", exc_info=True)
        return False


def parse_clipboard_data():
    """
    Parse the data from clipboard and create a DataFrame.
    
    Returns:
        pd.DataFrame or None: Parsed DataFrame with required columns
    """
    try:
        clipboard_content = pyperclip.paste()
        
        if not clipboard_content:
            logger.error("Clipboard is empty")
            return None
        
        logger.info(f"Clipboard content length: {len(clipboard_content)} characters")
        
        # Parse tab-separated data
        df = pd.read_csv(io.StringIO(clipboard_content), sep='\t', header=None)
        logger.info(f"Parsed DataFrame shape: {df.shape}")
        
        # Ensure we have at least 5 columns
        if df.shape[1] < 5:
            logger.error(f"Expected at least 5 columns, got {df.shape[1]}")
            return None
        
        # Take first 5 columns and rename them
        df = df.iloc[:, :5].copy()
        df.columns = EXPECTED_COLUMNS
        
        logger.info("DataFrame columns assigned successfully")
        logger.debug(f"First few rows:\n{df.head()}")
        
        return df
        
    except Exception as e:
        logger.error(f"Error parsing clipboard data: {e}", exc_info=True)
        return None


def transform_dataframe(df):
    """
    Apply required transformations to the DataFrame.
    
    Transformations:
    - Notes and Trade Amount: unchanged
    - Price: Convert XXX-YYY format to XXX + YYY/32
    - DV01 and DV01 Gamma: Remove commas and divide by 1000
    
    Args:
        df: Input DataFrame
        
    Returns:
        pd.DataFrame: Transformed DataFrame
    """
    try:
        df_transformed = df.copy()
        
        # Notes and Trade Amount columns - keep unchanged
        logger.info("Keeping Notes and Trade Amount columns unchanged")
        
        # Process Price column
        if 'Price' in df_transformed.columns:
            logger.info("Processing Price column...")
            df_transformed['Price'] = df_transformed['Price'].apply(parse_futures_price)
        
        # Process DV01 column
        if 'DV01' in df_transformed.columns:
            logger.info("Processing DV01 column...")
            # First clean the numeric values (remove commas)
            df_transformed['DV01'] = df_transformed['DV01'].apply(clean_numeric_value)
            # Then divide by 1000
            df_transformed['DV01'] = df_transformed['DV01'] / 1000.0
            
        # Process DV01 Gamma column
        if 'DV01 Gamma' in df_transformed.columns:
            logger.info("Processing DV01 Gamma column...")
            # First clean the numeric values (remove commas)
            df_transformed['DV01 Gamma'] = df_transformed['DV01 Gamma'].apply(clean_numeric_value)
            # Then divide by 1000
            df_transformed['DV01 Gamma'] = df_transformed['DV01 Gamma'] / 1000.0
        
        logger.info("Data transformation completed")
        logger.debug(f"Transformed data sample:\n{df_transformed.head()}")
        
        return df_transformed
        
    except Exception as e:
        logger.error(f"Error during data transformation: {e}", exc_info=True)
        return df


def write_to_excel(df, file_path=EXCEL_FILE_PATH):
    """
    Write DataFrame to Excel file at specified location.
    
    Args:
        df: DataFrame to write
        file_path: Path to Excel file
        
    Returns:
        bool: Success status
    """
    try:
        logger.info(f"Opening Excel file: {file_path}")
        
        # Check if file exists
        if not os.path.exists(file_path):
            logger.error(f"Excel file not found at: {file_path}")
            return False
        
        # Load workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Get or create target sheet
        if TARGET_SHEET in workbook.sheetnames:
            sheet = workbook[TARGET_SHEET]
            logger.info(f"Using existing sheet: {TARGET_SHEET}")
        else:
            sheet = workbook.create_sheet(TARGET_SHEET)
            logger.info(f"Created new sheet: {TARGET_SHEET}")
        
        # Clear existing data starting from target cell
        target_row = int(TARGET_CELL[1:])  # Extract row number from "A2"
        target_col = ord(TARGET_CELL[0]) - ord('A') + 1  # Convert 'A' to column 1
        
        # Write DataFrame starting at target cell
        logger.info(f"Writing DataFrame to {TARGET_CELL}...")
        
        # Get column indices for DV01 and DV01 Gamma
        dv01_col_idx = df.columns.get_loc('DV01') if 'DV01' in df.columns else -1
        dv01_gamma_col_idx = df.columns.get_loc('DV01 Gamma') if 'DV01 Gamma' in df.columns else -1
        
        # Write data without headers (assuming headers already exist in row 1)
        for r_idx, row in enumerate(df.values):
            for c_idx, value in enumerate(row):
                cell = sheet.cell(
                    row=target_row + r_idx,
                    column=target_col + c_idx,
                    value=value
                )
                
                # Apply number formatting to DV01 and DV01 Gamma columns to preserve decimals
                if c_idx == dv01_col_idx or c_idx == dv01_gamma_col_idx:
                    # Format as number with 3 decimal places
                    cell.number_format = '0.000'
        
        # Save workbook
        workbook.save(file_path)
        workbook.close()
        
        logger.info(f"Successfully wrote {len(df)} rows to Excel")
        return True
        
    except Exception as e:
        logger.error(f"Error writing to Excel: {e}", exc_info=True)
        if 'workbook' in locals():
            workbook.close()
        return False


def main():
    """
    Main execution function for the Pricing Monkey to Excel automation.
    """
    logger.info("=== Starting PmToExcel Automation ===")
    
    try:
        # Step 1: Browser automation
        logger.info("Step 1: Browser automation...")
        if not browser_automation():
            logger.error("Browser automation failed. Exiting.")
            return False
        
        # Small delay to ensure clipboard is ready
        time.sleep(0.5)
        
        # Step 2: Parse clipboard data
        logger.info("Step 2: Parsing clipboard data...")
        df = parse_clipboard_data()
        if df is None:
            logger.error("Failed to parse clipboard data. Exiting.")
            return False
        
        # Step 3: Transform data
        logger.info("Step 3: Transforming data...")
        df_transformed = transform_dataframe(df)
        
        # Step 4: Write to Excel
        logger.info("Step 4: Writing to Excel...")
        if not write_to_excel(df_transformed):
            logger.error("Failed to write to Excel. Exiting.")
            return False
        
        logger.info("=== PmToExcel Automation Completed Successfully ===")
        return True
        
    except Exception as e:
        logger.error(f"Unexpected error in main execution: {e}", exc_info=True)
        return False


if __name__ == "__main__":
    success = main()
    
    if success:
        logger.info("Script completed successfully")
        sys.exit(0)
    else:
        logger.error("Script failed")
        sys.exit(1) 