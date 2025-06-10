import openpyxl
from pathlib import Path

print("Test Excel Reader Starting...")

excel_file = Path('actantpnl.xlsx')
print(f"Current working directory: {Path.cwd()}")
print(f"Excel file exists: {excel_file.exists()}")

try:
    wb = openpyxl.load_workbook(excel_file)
    print(f"Workbook loaded successfully!")
    print(f"Sheet names: {wb.sheetnames}")
    
    ws = wb.active
    print(f"Active sheet: {ws.title}")
    print(f"Max row: {ws.max_row}, Max column: {ws.max_column}")
    
    # Try to read a few cells
    print("\nSample cells:")
    for row in range(1, min(5, ws.max_row + 1)):
        for col in range(1, min(5, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            print(f"  [{row},{col}] = {cell.value}")
            
except Exception as e:
    print(f"Error: {type(e).__name__}: {e}")
    import traceback
    traceback.print_exc()

print("\nTest complete.") 